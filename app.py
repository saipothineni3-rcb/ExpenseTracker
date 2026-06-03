from flask_mail import Mail, Message
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash


from openpyxl import Workbook
from reportlab.pdfgen import canvas

import os
from werkzeug.utils import secure_filename
print("Current Folder:", os.getcwd())
print("Templates Exists:", os.path.exists("templates"))

app = Flask(__name__)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'

app.config['MAIL_PORT'] = 587

app.config['MAIL_USE_TLS'] = True

app.config['MAIL_USERNAME'] = 'saipothineni3@gmail.com'

app.config['MAIL_PASSWORD'] = 'ubklguluhbtzlayr'

mail = Mail(app)
UPLOAD_FOLDER = 'static/receipts'

app.config[
    'UPLOAD_FOLDER'
] = UPLOAD_FOLDER

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)

app.config['SECRET_KEY'] = 'expense_tracker_secret'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'

db = SQLAlchemy(app)

# ======================
# User Table
# ======================
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    name = db.Column(db.String(100))

    email = db.Column(db.String(100), unique=True)

    phone = db.Column(db.String(15))

    password = db.Column(db.String(100))

# ======================
# Expense Table
# ======================

class Expense(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    category = db.Column(
        db.String(100)
    )

    amount = db.Column(
        db.Float
    )

    description = db.Column(
        db.String(200)
    )

    date = db.Column(
        db.String(20)
    )

    receipt = db.Column(
        db.String(200)
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey('user.id')
    )
class Budget(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    monthly_budget = db.Column(db.Float)
class Goal(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    target_amount = db.Column(
        db.Float
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey('user.id')
    )
# ======================
# Goal Table
# ======================



@app.route('/set_budget', methods=['GET', 'POST'])
def set_budget():

    if request.method == 'POST':

        amount = float(request.form['budget'])

        budget = Budget(
            monthly_budget=amount
        )

        db.session.add(budget)
        db.session.commit()

        return redirect('/dashboard')

    return render_template('set_budget.html')


# ======================
# Create Database
# ======================
with app.app_context():
    db.create_all()


# ======================
# Home Page
# ======================
@app.route('/')
def home():
    return redirect('/register')


# ======================
# Register Page
# ======================
@app.route('/register', methods=['GET', 'POST'])
def register():

    if request.method == 'POST':

        name = request.form['name']

        email = request.form['email']

        phone = request.form['phone']

        password = request.form['password']

        existing_user = User.query.filter_by(
            email=email
        ).first()

        if existing_user:
            return redirect('/login')

        hashed_password = generate_password_hash(
            password
        )

        user = User(
            name=name,
            email=email,
            phone=phone,
            password=hashed_password
        )

        db.session.add(user)
        db.session.commit()

        return redirect('/login')

    return render_template(
        'register.html'
    )
#=======================
# Login Page
# ======================

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']

        user = User.query.filter_by(
            email=email
        ).first()

        if user and check_password_hash(user.password, password):

            session['user_id'] = user.id
            session['user_name'] = user.name

            return redirect('/dashboard')

        else:
            return "Invalid Email or Password"

    return render_template('login.html')

# ======================
# Dashboard
# ======================
@app.route('/dashboard')
def dashboard():

    if 'user_id' not in session:
        return redirect('/login')

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).all()

    total_amount = sum(
        expense.amount
        for expense in expenses
    )

    total_transactions = len(expenses)

    budget = Budget.query.order_by(
        Budget.id.desc()
    ).first()

    budget_amount = 0
    remaining_budget = 0

    if budget:

        budget_amount = budget.monthly_budget

        remaining_budget = (
            budget_amount - total_amount
        )
        # ======================
    # Savings Goal Tracker
    # ======================

    goal = Goal.query.filter_by(
        user_id=session['user_id']
    ).order_by(
        Goal.id.desc()
    ).first()

    goal_amount = 0

    goal_progress = 0

    if goal:

        goal_amount = goal.target_amount

        if goal_amount > 0:

            goal_progress = min(

                round(
                    (
                        remaining_budget /
                        goal_amount
                    ) * 100,
                    1
                ),

                100
            )
    # ======================
    # Notifications
    # ======================

    notifications = []

    if budget_amount > 0:

        percentage = (
            total_amount / budget_amount
        ) * 100

        if percentage >= 100:

            notifications.append(
                "🚨 Budget exceeded."
            )

        elif percentage >= 80:

            notifications.append(
                "⚠ More than 80% budget used."
            )

        elif percentage <= 50:

            notifications.append(
                "✅ Spending is under control."
            )

    # ======================
    # Financial Health Score
    # ======================

    score = 100

    if budget_amount > 0:

        percentage = (
            total_amount / budget_amount
        ) * 100

        if percentage > 100:

            score = 30

        elif percentage > 90:

            score = 50

        elif percentage > 80:

            score = 70

        elif percentage > 60:

            score = 85

        else:

            score = 100

    else:

        score = 50

    # ======================
    # Health Status
    # ======================

    if score >= 90:

        health_status = "🏆 Excellent"

    elif score >= 70:

        health_status = "✅ Good"

    elif score >= 50:

        health_status = "⚠ Average"

    else:

        health_status = "🚨 Poor"

    # ======================
    # Achievement Badge
    # ======================

    badge = "🎯 Beginner"

    if budget_amount > 0:

        savings_percent = (
            remaining_budget /
            budget_amount
        ) * 100

        if savings_percent >= 50:

            badge = "🥇 Financial Master"

        elif savings_percent >= 25:

            badge = "🥈 Smart Saver"

        elif savings_percent >= 10:

            badge = "🥉 Budget Saver"

    # ======================
    # Budget Message
    # ======================

    if remaining_budget < 0:

        budget_message = (
            f"🚨 Budget Exceeded by ₹{abs(remaining_budget)}"
        )

        budget_color = "danger"

    else:

        budget_message = (
            f"✅ ₹{remaining_budget} Budget Remaining"
        )

        budget_color = "success"

    # ======================
    # Daily Expense Summary
    # ======================

    daily_expenses = {}

    for expense in expenses:

        if expense.date in daily_expenses:

            daily_expenses[
                expense.date
            ] += expense.amount

        else:

            daily_expenses[
                expense.date
            ] = expense.amount
        # ======================
    # Expense Streak
    # ======================

    streak = len(
        set(
            expense.date
            for expense in expenses
        )
    )
    # ======================
    # Recent Expenses
    # ======================

    recent_expenses = Expense.query\
        .filter_by(
            user_id=session['user_id']
        )\
        .order_by(
            Expense.id.desc()
        )\
        .limit(5)\
        .all()

    # ======================
    # Category Totals
    # ======================

    category_totals = {}

    for expense in expenses:

        if expense.category in category_totals:

            category_totals[
                expense.category
            ] += expense.amount

        else:

            category_totals[
                expense.category
            ] = expense.amount
        # ======================
    # AI Spending Insights
    # ======================

    insights = []

    if category_totals:

        top_category = max(
            category_totals,
            key=category_totals.get
        )

        insights.append(
            f"💡 {top_category} is your highest expense category."
        )

    if budget_amount > 0:

        usage = round(
            (total_amount / budget_amount) * 100,
            1
        )

        insights.append(
            f"📊 You have used {usage}% of your budget."
        )

        if usage >= 100:

            insights.append(
                "🚨 Budget exceeded. Reduce spending immediately."
            )

        elif usage >= 80:

            insights.append(
                "⚠ Budget usage is very high."
            )

        else:

            insights.append(
                "✅ Spending is under control."
            )

    if "Shopping" in category_totals:

        insights.append(
            "🛍 Consider reducing Shopping expenses to save more money."
        )
    # ======================
    # Dashboard Statistics
    # ======================

    largest_expense = 0

    if expenses:

        largest_expense = max(
            expense.amount
            for expense in expenses
        )

    average_expense = 0

    if total_transactions > 0:

        average_expense = round(
            total_amount /
            total_transactions,
            2
        )

    days_active = len(
        set(
            expense.date
            for expense in expenses
        )
    )

    most_used_category = "N/A"

    if category_totals:

        most_used_category = max(
            category_totals,
            key=category_totals.get
        )

    return render_template(

        'dashboard.html',

        total_amount=total_amount,

        total_transactions=total_transactions,

        budget_amount=budget_amount,

        remaining_budget=remaining_budget,

        recent_expenses=recent_expenses,

        category_totals=category_totals,

        notifications=notifications,

        budget_message=budget_message,

        budget_color=budget_color,

        score=score,

        health_status=health_status,

        badge=badge,

        daily_expenses=daily_expenses,

        largest_expense=largest_expense,

        average_expense=average_expense,

        days_active=days_active,

        most_used_category=most_used_category,

        user_name=session['user_name'],
        
        insights=insights,
        goal_amount=goal_amount,

        goal_progress=goal_progress,
        
        streak=streak,
    )
# ======================
# Add Expense
# ======================
@app.route(
    '/add_expense',
    methods=['GET', 'POST']
)
def add_expense():

    if request.method == 'POST':

        category = request.form[
            'category'
        ]

        amount = float(
            request.form['amount']
        )

        description = request.form[
            'description'
        ]

        receipt_file = request.files[
            'receipt'
        ]

        filename = ''

        if receipt_file.filename != '':

            filename = secure_filename(
                receipt_file.filename
            )

            receipt_file.save(

                os.path.join(
                    app.config[
                        'UPLOAD_FOLDER'
                    ],
                    filename
                )
            )

        expense = Expense(

            category=category,

            amount=amount,

            description=description,

            date=str(
                datetime.now().date()
            ),

            receipt=filename,

            user_id=session[
                'user_id'
            ]
        )

        db.session.add(
            expense
        )

        db.session.commit()

        return redirect(
            '/view_expenses'
        )

    return render_template(
        'add_expense.html'
    )

# ======================
# View Expenses
# ======================
from datetime import datetime

@app.route('/view_expenses')
def view_expenses():

    filter_type = request.args.get('filter')

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).all()

    today = datetime.now().date()

    filtered = []

    for expense in expenses:

        try:

            expense_date = datetime.strptime(
                expense.date,
                "%Y-%m-%d"
            ).date()

            if filter_type == "today":

                if expense_date == today:
                    filtered.append(expense)

            elif filter_type == "month":

                if (
                    expense_date.month ==
                    today.month
                ):
                    filtered.append(expense)

            else:

                filtered.append(expense)

        except:

            filtered.append(expense)

    return render_template(
        'view_expenses.html',
        expenses=filtered
    )

@app.route('/delete_expense/<int:id>')
def delete_expense(id):

    expense = Expense.query.get(id)

    db.session.delete(expense)

    db.session.commit()

    return redirect('/view_expenses')
@app.route('/analytics')
def analytics():

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).all()

    categories = []
    amounts = []

    monthly_data = {}

    for expense in expenses:

        categories.append(expense.category)
        amounts.append(expense.amount)

        if expense.date:

            month = expense.date[:7]

            if month in monthly_data:

                monthly_data[month] += expense.amount

            else:

                monthly_data[month] = expense.amount

    months = list(monthly_data.keys())
    monthly_amounts = list(monthly_data.values())

    return render_template(
        'analytics.html',
        categories=categories,
        amounts=amounts,
        months=months,
        monthly_amounts=monthly_amounts
    )
@app.route('/edit_expense/<int:id>', methods=['GET', 'POST'])
def edit_expense(id):

    expense = Expense.query.get(id)

    if request.method == 'POST':

        expense.category = request.form['category']
        expense.amount = float(request.form['amount'])
        expense.description = request.form['description']
        expense.date = request.form['date']

        db.session.commit()

        return redirect('/view_expenses')

    return render_template(
        'edit_expense.html',
        expense=expense
    )
@app.route('/logout')
def logout():

    session.clear()

    return redirect('/login')
@app.route('/export_excel')
def export_excel():

    expenses = Expense.query.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Expenses"

    ws.append([
        "ID",
        "Category",
        "Amount",
        "Description"
    ])

    for expense in expenses:

        ws.append([
            expense.id,
            expense.category,
            expense.amount,
            expense.description
        ])

    wb.save("expenses.xlsx")

    return send_file(
        "expenses.xlsx",
        as_attachment=True
    )
@app.route('/export_pdf')
def export_pdf():

    expenses = Expense.query.all()

    pdf_file = "expenses_report.pdf"

    c = canvas.Canvas(pdf_file)

    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, 800, "Expense Report")

    y = 750

    for expense in expenses:

        text = (
            f"ID: {expense.id} | "
            f"Category: {expense.category} | "
            f"Amount: ₹{expense.amount} | "
            f"Description: {expense.description}"
        )

        c.drawString(50, y, text)

        y -= 25

        if y < 50:
            c.showPage()
            y = 750

    c.save()

    return send_file(
        pdf_file,
        as_attachment=True
    )
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():

    if request.method == 'POST':

        email = request.form['email']

        last4 = request.form['last4']

        user = User.query.filter_by(
            email=email
        ).first()

        if user:

            if user.phone[-4:] == last4:

                session['reset_user'] = user.id

                return redirect(
                    '/reset_password'
                )

            else:

                return render_template(
                    'error.html',
                    title='Verification Failed',
                    message='Last 4 digits of phone number are incorrect.'
                )

        else:

            return render_template(
                'error.html',
                title='User Not Found',
                message='No account exists with this email address.'
            )

    return render_template(
        'forgot_password.html'
    )
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():

    if 'reset_user' not in session:

        return redirect('/login')
    if request.method == 'POST':

        new_password = request.form['password']

        user = User.query.get(
            session['reset_user']
        )

        user.password = generate_password_hash(
            new_password
        )

        db.session.commit()

        session.pop('reset_user')

        return redirect('/login')

    return render_template(
        'reset_password.html'
    )
@app.route('/check_users')
def check_users():

    users = User.query.all()

    result = ""

    for user in users:

        result += f"""
        Name: {user.name}<br>
        Email: {user.email}<br>
        Phone: {user.phone}<br><br>
        """

    return result

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():

    if 'user_id' not in session:
        return redirect('/login')

    if request.method == 'POST':

        current_password = request.form['current_password']

        new_password = request.form['new_password']

        user = User.query.get(
            session['user_id']
        )

        if check_password_hash(
            user.password,
            current_password
        ):

            user.password = generate_password_hash(
                new_password
            )

            db.session.commit()

            return render_template(
                'success.html'
            )

        else:

            return render_template(
                'error.html',
                title='Incorrect Password',
                message='Current password is incorrect.'
            )

    return render_template(
        'change_password.html'
    )
@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():

    if 'user_id' not in session:
        return redirect('/login')

    if request.method == 'POST':

        file = request.files['file']

        wb = load_workbook(file)

        ws = wb.active

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            print("ROW =", row)

            try:

                expense = Expense(

                    category=str(row[0]),

                    amount=float(row[1]),

                    description=str(row[2]),

                    date=str(row[3]),

                    user_id=session['user_id']
                )

                db.session.add(expense)

            except Exception as e:

                print("ERROR:", e)
                print("PROBLEM ROW:", row)

        db.session.commit()

        return redirect('/view_expenses')

    return render_template(
        'import_excel.html'
    )
@app.route('/set_goal', methods=['GET', 'POST'])
def set_goal():

    if request.method == 'POST':

        amount = float(
            request.form['goal']
        )

        goal = Goal(
            amount=amount
        )

        db.session.add(goal)

        db.session.commit()

        return redirect(
            '/dashboard'
        )

    return render_template(
        'set_goal.html'
    )
@app.route('/search_expense')
def search_expense():

    keyword = request.args.get(
        'keyword'
    )

    expenses = Expense.query.filter(
        Expense.category.contains(
            keyword
        )
    ).all()

    return render_template(
        'search_results.html',
        expenses=expenses,
        keyword=keyword
    )
@app.route('/send_report')
def send_report():

    if 'user_id' not in session:
        return redirect('/login')

    user = User.query.get(
        session['user_id']
    )

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).all()

    total_amount = sum(
        expense.amount
        for expense in expenses
    )

    total_transactions = len(
        expenses
    )

    msg = Message(

        "Expense Tracker Report",

        sender=app.config[
            'MAIL_USERNAME'
        ],

        recipients=[
            user.email
        ]
    )

    msg.body = f"""
Hello {user.name}

Expense Summary

Total Expenses: ₹{total_amount}

Transactions: {total_transactions}

Thank you for using Expense Tracker.
"""

    mail.send(msg)

    return render_template(
        'success.html',
        title='Email Sent',
        message='Report sent successfully.'
    )
from flask import send_file
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import (
    getSampleStyleSheet
)

@app.route('/monthly_report')
def monthly_report():

    if 'user_id' not in session:
        return redirect('/login')

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).all()

    total_amount = sum(
        expense.amount
        for expense in expenses
    )

    total_transactions = len(
        expenses
    )

    filename = "monthly_report.pdf"

    pdf = SimpleDocTemplate(
        filename
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "Expense Tracker Monthly Report",
            styles['Title']
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    elements.append(
        Paragraph(
            f"User: {session['user_name']}",
            styles['Normal']
        )
    )

    elements.append(
        Paragraph(
            f"Total Expenses: ₹{total_amount}",
            styles['Normal']
        )
    )

    elements.append(
        Paragraph(
            f"Transactions: {total_transactions}",
            styles['Normal']
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    for expense in expenses:

        elements.append(

            Paragraph(

                f"{expense.date} | "
                f"{expense.category} | "
                f"₹{expense.amount}",

                styles['Normal']
            )

        )

    pdf.build(
        elements
    )

    return send_file(
        filename,
        as_attachment=True
    )
@app.route('/profile', methods=['GET', 'POST'])
def profile():

    if 'user_id' not in session:
        return redirect('/login')

    user = User.query.get(
        session['user_id']
    )

    if request.method == 'POST':

        user.name = request.form['name']

        user.email = request.form['email']

        user.phone = request.form['phone']

        db.session.commit()

        session['user_name'] = user.name

        return redirect('/profile')

    return render_template(
        'profile.html',
        user=user
    )
# ======================
# Run Application
# ======================
if __name__ == '__main__':
    app.run(debug=True)